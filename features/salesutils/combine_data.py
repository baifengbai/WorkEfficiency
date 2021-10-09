# -*- coding: utf-8 -*-
# @Time    : 9/27/2021 8:23 PM
# @Author  : Chris.Wang
# @Site    : 
# @File    : combine_data.py
# @Software: PyCharm
# @Description:
import copy
import datetime
import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, NamedStyle
from openpyxl.utils import get_column_letter

columns = ['date/time', 'settlement id', 'type', 'order id', 'sku', 'description',
           'quantity', 'marketplace', 'fulfilment', 'order city', 'order state',
           'order postal', 'tax collection model', 'product sales',
           'product sales tax', 'postage credits', 'shipping credits tax',
           'gift wrap credits', 'giftwrap credits tax', 'promotional rebates',
           'promotional rebates tax', 'marketplace withheld tax', 'selling fees',
           'fba fees', 'other transaction fees', 'other', 'total']

def ret_dataframe(fpath,sku_list,fdate_list,subs_list):
    dirpath, fp = os.path.split(fpath)
    filename, ext = os.path.splitext(fp)
    site = filename[-2:]
    eu_sites = ['nl','fr','de','es','it']
    print(f'载入: {fpath}')
    if site in eu_sites:
        decimal_ = ','
    else:
        decimal_ = '.'
    df = pd.read_csv(fpath,encoding='UTF-8',skiprows=7,decimal=decimal_)
    df.columns = columns
    # df = df.apply(lambda x: x.str.replace(',', '.'))
    # df.loc[:, 'product sales':'total'] = pd.to_numeric(df.loc[:, 'product sales':'total'])

    if site in eu_sites:
        df2 = df.loc[:,'product sales':'total']
        for key in df2.columns:
            df[key] = pd.to_numeric(df[key],errors='coerce')
        # df.loc[:'product sales':'total'] = pd.to_numeric(df.loc[:, 'product sales':'total'],errors='coerce')
        df.loc[:, 'product sales':'total'] = df.loc[:, 'product sales':'total'].mul(7.5/8.5, axis=1)
        print(f'站点为：{site}, 实行货币单位换算。')
        for index,row in df.iterrows():
            if '.' in row['date/time']:
                # change date
                clist = re.findall('([0-9]+)\.([0-9]+)\.(.+)', row['date/time'])
                if len(clist) > 0 and len(clist[0]) == 3:
                    clist_ = '/'.join([clist[0][1], clist[0][0], clist[0][2]])
                    df.loc[index, 'date/time'] = clist_
            for fdate in fdate_list:
                slice = row['date/time'].split(' ')
                if slice[1].lower() in fdate:
                    df.loc[index,'date/time'] = df.loc[index,'date/time'].replace(slice[1].lower(),fdate[0])
                    break
            for subs in subs_list:
                if row['type'] in subs:
                    df.loc[index,'type'] = df.loc[index,'type'].replace(row['type'],subs[0])
                    break

    df['date/time'] = pd.to_datetime(df['date/time'])
    # Remove timezone from columns
    df['date/time'] = df['date/time'].dt.tz_localize(None)
    df['项目'] = ''
    for index, row in df.iterrows():
        for i, sku in enumerate(sku_list):
            if sku_list[i][0] == row['sku']:
                df.loc[index, '项目'] = sku_list[i][1]
    return df

# main def
def read_fileset():
    folder_path = input('请输入目录路径: ').strip('\"').strip()
    time_start = datetime.datetime.now()
    result = os.walk(folder_path)
    result2 = []
    sku_proj = None
    fdate_path = None
    subs_path = None
    flag_sku = True
    flag_fdate = True
    flag_subs = True
    for path,dirs,files in result:
        for file in files:
            _,ext = os.path.splitext(file)
            if ext.lower() in ['.csv']:
                result2.append((path,file))
            if flag_sku and (file == 'SKU-PROJECT.xlsx'):
                sku_proj = os.path.join(path,file)
                flag_sku = False
            if flag_fdate and (file == 'DATE-FORMAT.xlsx'):
                fdate_path = os.path.join(path,file)
                flag_fdate = False
            if flag_subs and (file == 'TYPE-MATCH.xlsx'):
                subs_path = os.path.join(path,file)
                flag_subs = False
    df = pd.DataFrame(columns=columns)
    print('载入SKU项目对应表。')
    sku_df = pd.read_excel(sku_proj)
    sku_list = list(zip(sku_df['SKU'],sku_df['项目']))
    print('载入日期对应表。')
    fdate_df = pd.read_excel(fdate_path)
    fdate_list = [fdate_df[c].tolist() for c in fdate_df.columns]

    print('载入类型对应表。')
    subs_df = pd.read_excel(subs_path)
    subs_list = [subs_df[c].tolist() for c in subs_df.columns]

    for path,file in result2:
        filepath = os.path.join(path,file)
        df_temp = ret_dataframe(filepath,sku_list,fdate_list,subs_list)
        df = df.append(df_temp,ignore_index=True)
        # print(df)
    df = df.drop(['tax collection model'], axis=1)
    df = df.drop(['marketplace withheld tax'], axis=1)
    save_path = os.path.join(folder_path,'店铺订单数据汇总.xlsx')

    # escape unicode character
    df = df.applymap(lambda x: x.encode('unicode_escape').
                                   decode('utf-8') if isinstance(x, str) else x)
    df.to_excel(save_path,engine='openpyxl')

    print('处理汇总文件。')
    wb = load_workbook(save_path)
    ws1 = wb.active
    ws1.delete_cols(1,1)

    # 生成表头名称字典
    col_names = {}
    current = 1
    for col in ws1.iter_cols(1, ws1.max_column):
        col_names[col[0].value] = get_column_letter(current)
        current += 1
    def remove_formatting(ws):
        # ws is not the worksheet name, but the worksheet object
        for row in ws.iter_rows():
            for cell in row:
                cell.style = 'Normal'
    remove_formatting(ws1)

    # 格式化时间
    date_style = NamedStyle(name='c_datetime', number_format="YYYY-MM-DD HH:mm:ss")
    for cell in ws1[col_names['date/time']][1:]:
        cell.style = date_style

    # 设定字体
    for col in ws1.columns:
        for cell in col:
            cell.font = Font(name="Calibri")
            cell.alignment = Alignment(horizontal='left')

    # 标题设定加粗
    for cell in ws1[1:1]:
        font1 = copy.copy(cell.font)
        font1.bold = True
        cell.font = font1
    ws1.column_dimensions[col_names['date/time']].width = 19


    wb.save(save_path)
    print('汇总文件保存。')
    time_end = datetime.datetime.now()
    timediff = time_end - time_start
    print(f"耗时: {timediff.seconds}s.")



if __name__ == '__main__':
    read_fileset()


