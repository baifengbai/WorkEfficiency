import os
import shutil

import filetype
import PIL.Image
from PIL import ImageOps
from PIL.JpegImagePlugin import JpegImageFile

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def resize_im(im: JpegImageFile, size: tuple, fillcolor: tuple)-> JpegImageFile:
    # 图像处理函数
    # size[0] is height, size[1] is width
    (width, height) = size

    required_ratio = height / width
    im_width, im_height = im.size
    im_ratio = im_height / im_width
    # 太高了，需要加宽
    if im_ratio >= required_ratio:
        # 原图高度为最大限制
        resize_width, resize_height = int(height / im_ratio), height
        source_resized = im.resize((resize_width, resize_height))
        append_width_l = int((width - resize_width) / 2)
        append_width_r = width - resize_width - append_width_l
        # left, top, right, bottom = _border(border)
        im_dest = ImageOps.expand(
            source_resized,
            border=(append_width_l, 0, append_width_r, 0),
            fill=fillcolor)
    else:
        # 原图宽度为最大限制
        resize_width, resize_height = width, int(width * im_ratio)
        source_resized = im.resize((resize_width, resize_height))
        append_height_t = int((height - resize_height) / 2)
        append_height_b = height - resize_height - append_height_t
        im_dest = ImageOps.expand(
            source_resized,
            border=(0, append_height_t, 0, append_height_b),
            fill=fillcolor)

    return im_dest


def pic_to_excel():

    # excel path
    excel_path = input('输入excel文件路径：').strip('\"').strip()
    # pic resource path
    # default
    res_path = input('输入图片资源文件夹路径：').strip('\"').strip() or os.path.dirname(excel_path)
    resfile_list = os.listdir(res_path)
    resfile_list_ = []
    for res in resfile_list:
        filename_, ext_ = os.path.splitext(res)
        if ext_.lower() in ['.jpeg', '.jpg', '.png']:
            resfile_list_.append((res_path,filename_,ext_))
    resfile_list = resfile_list_.copy()

    # open excel
    wb = load_workbook(excel_path)
    ws = wb.active
    # ws['A1'].value = "code"
    # ws['A1'].alignment = Alignment(horizontal='center')
    # ws['A1'].font = Font(bold=True)
    # set font

    # 生成表头名称字典
    col_names = {}
    current = 1
    for col in ws.iter_cols(1, ws.max_column):
        col_names[str(col[0].value).lower()] = get_column_letter(current)
        current += 1

    # 设定字体
    for col in ws.columns:
        for cell in col:
            cell.font = Font(name="Calibri")

    # 文本居中
    # for col in ws.columns:
    #     for cell in col:
    #         cell.alignment = Alignment(wrap_text=True, vertical='center')

    # 设定列表宽度
    col_width = 12.5
    ws.column_dimensions[col_names['im']].width = col_width  # img

    # 添加图片
    sheetfile_list = ws[col_names['fn']]
    lencode = len(sheetfile_list)

    print('开始添加图片')
    fillcolor = (255, 255, 255)
    height_ = 300
    width_ = 300
    set_size = (height_, width_)
    temppic = 'temppic'
    if not os.path.exists(os.path.join(res_path,temppic)):
        os.mkdir(os.path.join(res_path,temppic))
    for x in range(1, lencode):
        filename = sheetfile_list[x].value
        index_ = None
        for i, file in enumerate(resfile_list):
            if file[1] == filename:
                index_ = i
                break
        if index_ is not None:
            img_path = os.path.join(res_path,''.join(resfile_list[index_][1:]))
            img = PIL.Image.open(img_path)
            img = resize_im(img, size=set_size, fillcolor=fillcolor)
            temp_img_path = os.path.join(res_path,temppic,''.join(resfile_list[index_][1:]))
            img.save(temp_img_path)
            im = Image(temp_img_path)
            width = im.width
            height = im.height
            im.width =col_width * 7
            im.height = im.width / width * height
            ws.row_dimensions[x + 1].height = col_width * height / width * 6
            im.anchor = f'B{x + 1}'
            print(f'添加图片：{filename}')
            ws.add_image(im)

    wb.save(excel_path)
    shutil.rmtree(os.path.join(res_path,temppic))


if __name__ == '__main__':
    pic_to_excel()