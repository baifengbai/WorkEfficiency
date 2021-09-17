import os
import shutil
import tempfile
import traceback
import typing

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import filetype

import os

from pdf2image import convert_from_path

import PIL.Image
import pytesseract

from pyzbar.pyzbar import decode

import PyPDF2

import pandas as pd

features = {}
features_id = {}
features_description = {}


def register_feature(description="暂无描述"):
    def wrap(func):
        if func.__name__ not in features.keys():
            feature_name = func.__name__
            features[feature_name] = func
            features_description[feature_name] = description
        return func

    return wrap


def run_feature(func):
    print(f"运行功能：{func.__name__}")
    return func()


def show_description():
    for n, (k, v) in enumerate(features.items()):
        features_id[n] = k
        print(f'{n}: {k}')
        print(features_description[k])

@register_feature(description="""用于显示所有的功能
""")
def show_features():
    print('目前以下功能可用：')
    for n, (k, v) in enumerate(features.items()):
        features_id[n] = k
        print(f'{n}: {k}', end=' ')
        print(f"({features_description[k].splitlines()[0]})")

# -------------------------------------------------------------------


def _read_text_pdf():
    fpath = input('请输入文件路径：').strip('\"')
    with open(fpath, 'rb') as fp:
        pdfFileObject = fp
        pdfReader = PyPDF2.PdfFileReader(pdfFileObject, strict=True)
        print(" No. Of Pages: ", pdfReader.numPages)
        pageObject = pdfReader.getPage(0)
        return pageObject.extractText()

def _decode_img(img: PIL.Image,barcode_mode=True,text_mode=True):
    if barcode_mode:
        detectedBarcodes = decode(img)
        # If not detected then print the message
        barcode_ = []
        if not detectedBarcodes:
            print("Barcode Not Detected or your barcode is blank/corrupted!")
        else:
            # Traveres through all the detected barcodes in image
            for barcode in detectedBarcodes:

                # # Locate the barcode position in image
                # (x, y, w, h) = barcode.rect
                #
                # # Put the rectangle in image using
                # # cv2 to heighlight the barcode
                # cv2.rectangle(img, (x - 10, y - 10),
                #               (x + w + 10, y + h + 10),
                #               (255, 0, 0), 2)
                if barcode.data != "":
                    # Print the barcode data
                    barcode_.append(barcode.data.decode('ascii'))
                    # print(barcode.type)
    else:
        barcode_ = []
    if text_mode:
        s1 = pytesseract.image_to_string(img).strip()
        while True:
            s1 = s1.replace('\n\n', '\n')
            if "\n\n" not in s1:
                break
        text = s1.splitlines()
    else:
        text = ''
    return barcode_, text

# 对图片进行分析，获得条码数据和文字数据
def _read_bartag(fpath) -> (list,list):
    kind = filetype.guess(fpath)
    img = typing.Any
    if kind.extension == "pdf":
        images_from_path = convert_from_path(fpath, dpi=600)
        for im in images_from_path:
            img = im
            break
    elif kind.extension in ['png', 'jpg', 'jpeg']:
        img = PIL.Image.open(fpath)
    else:
        exit('Unsupported file type. ')
    return _decode_img(img)

def _splitter_(splitter: list, pdfReader: PyPDF2.PdfFileReader, fpath, pages_data=None):
    root, fname = os.path.split(fpath)
    count = 0
    for spl in splitter:
        count += 1
        pdf_writer = PyPDF2.PdfFileWriter()
        for i in range(spl[0], spl[1] + 1):
            current_page = pdfReader.getPage(i)
            pdf_writer.addPage(current_page)
        if pages_data is None:
            prefix = fname
        else:
            lenx = len(pages_data[spl[0]][0])
            prefix = pages_data[spl[0]][0][lenx-1]
        output_file = f"{prefix}-{count}-{spl[1] + 1 - spl[0]}-pages.pdf"
        with open(os.path.join(root, output_file), 'wb') as outfp:
            pdf_writer.write(outfp)
        print(f'分割文件：{output_file}')

@register_feature(description="""将pdf保存为图片
""")
def pdf_save2img():
    f_path = input('请输入处理文件夹：').strip('\"')
    result_path = input('请输入结果保存位置：').strip('\"')

    temp_file_list = os.listdir(f_path)
    file_list = []

    for file in temp_file_list:
        if ".pdf" in file:
            file_list.append(file)
    cot = len(file_list)
    print(f'You have {cot} pdf file(s).')
    n = 0
    temp_dir = os.path.join(f_path, 'temp-asgdafewa')
    if os.path.exists(temp_dir):
        shutil.rmtree(temp_dir)
    os.mkdir(temp_dir)
    for f in file_list:
        images_from_path = convert_from_path(os.path.join(f_path, f),output_folder=temp_dir, dpi=600,thread_count=8)
        name, extension = os.path.splitext(f)
        for i,im in enumerate(images_from_path):
            im.save(os.path.join(result_path, name + f"-{i}.jpg"))
        n += 1
        print(f"finished {n} in {cot}.")
    shutil.rmtree(temp_dir)

# excel 添加条码
@register_feature(description="""为excel表格添加条码图片
""")
def tag_to_excel():
    save_fpath = input("请输入文件路径：").strip('\"')
    resource_path = input('请输入素材文件夹(支持图片和单页pdf)：').strip('\"')
    res_filelist = os.listdir(resource_path)
    res_filelist_ = []
    for res in res_filelist:
        if os.path.isfile(os.path.join(resource_path,res)):
            res_filelist_.append(os.path.join(resource_path,res))
    res_filelist = res_filelist_.copy()
    lenres = len(res_filelist)
    wb = load_workbook(save_fpath)
    ws = wb.active
    # ws['A1'].value = "code"
    # ws['A1'].alignment = Alignment(horizontal='center')
    # ws['A1'].font = Font(bold=True)
    # set font

    # 生成表头名称字典
    col_names = {}
    current = 1
    for col in ws.iter_cols(1, ws.max_column):
        col_names[str(col[0].value).lower()] = get_column_letter(current)
        current += 1

    # 设定字体
    for col in ws.columns:
        for cell in col:
            cell.font = Font(name="Calibri")

    # 文本居中
    # for col in ws.columns:
    #     for cell in col:
    #         cell.alignment = Alignment(wrap_text=True, vertical='center')

    # 设定列表宽度
    col_width = 36
    ws.column_dimensions[col_names['tag']].width = col_width  # img

    # 添加图片
    code_list = ws[col_names['code']]
    lencode = len(code_list)

    def find_code_res(code_, filelist,lenfl):
        for i in range(lenfl):
            _, ext = os.path.splitext(filelist[i])
            if (code_ in filelist[i]) and (ext.lower() in ['.jpeg','.jpg','.png','.gif','.webp','.pdf']):
                return filelist[i]
            else:
                continue
        return None

    for x in range(1, lencode):
        img = None
        code = code_list[x].value
        target = find_code_res(code,res_filelist,lenres)
        if not target:
            continue
        else:
            _,ext = os.path.splitext(target)
            if ext.lower() == '.pdf':
                temp_dir = os.path.join(resource_path,'temp-asgdafewa')
                if os.path.exists(temp_dir):
                    shutil.rmtree(temp_dir)
                os.mkdir(temp_dir)
                images_from_path = convert_from_path(target, dpi=600,output_folder=temp_dir,thread_count=8)
                for im in images_from_path:
                    img = Image(im)
                    break
            elif ext.lower() in ['.jpeg','.jpg','.png','.gif','.webp']:
                img = Image(target)
            else:
                continue
            width = img.width
            height = img.height
            img.width = min(max(width * 0.45, col_width * 7), col_width * 7)
            img.height = img.width / width * height
            ws.row_dimensions[x + 1].height = col_width * height / width * 6
            img.anchor = f'B{x + 1}'
            print(f'添加图片：{code}')
            ws.add_image(img)
    wb.save(save_fpath)

# 验证fnsku与sku对应关系
@register_feature(description="""验证两组code对应关系匹配
""")
def fba_verify_barcode():
    vfilepath = input('请输入验证表格文件路径 (必须含有sku和fnsku) ：').strip('\"') # 必须含有sku和fnsku
    bardirpath = input('请输入条码文件夹路径：').strip('\"')
    vdf = pd.read_excel(vfilepath)
    vdf.columns = vdf.columns.str.lower()
    bar_list = os.listdir(bardirpath)
    bar_list_full = [os.path.join(bardirpath, barfile) for barfile in bar_list]
    print('识别条码数据...')
    count = 0
    lenf = len(bar_list_full)

    def _x(b,i,lenf):
        print(f'\r识别第{i}/{lenf}个文件.',end='')
        return _read_bartag(b)

    bar_data = [_x(b,n+1, lenf) for n,b in enumerate(bar_list_full)]
    print('\n开始数据验证...')
    flag = True
    for i in range(len(bar_list_full)):
        bc, s = bar_data[i]
        if (bc[0] in vdf['fnsku'].values) and (vdf.loc[vdf['fnsku'] == bc[0]]['sku'].iloc[0] in ' '.join(s)):
            pass
        else:
            name, ext = os.path.splitext(bar_list[i])
            print(f'数据不匹配: {bar_list[i]}')
            if flag:
                os.mkdir(os.path.join(bardirpath,'unmatch'))
            os.rename(bar_list_full[i], os.path.join(bardirpath,'unmatch',name + ext))
            flag = False
    if flag:
        print('所有条码数据正确。')
    else:
        print('错误条码已标注，请修正。')

#barcode renamer,文件名修改为条码数据,for pdf and jpg
@register_feature(description="""文件名修改为条码数据
""")
def read_barcode():
    while True:
        mode = input('选择模式(print or rename)：')
        if mode not in ['print','rename']:
            exit(f'模式不支持：{mode}')
        else:
            break

    path = input("输入识别文件夹路径：").strip('\"')
    fl = os.listdir(path)
    fl_tmp = []
    for f in fl:
        f_ = os.path.join(path,f)
        if os.path.isfile(f_) and filetype.guess(f_).extension in ['pdf','jpg','png']:
            fl_tmp.append(f)
    fl = fl_tmp.copy()
    result_path = os.path.join(path, 'result')
    if mode == 'rename':
        if os.path.exists(result_path):
            shutil.rmtree(result_path)
        os.mkdir(result_path)
    for f in fl:
        f_name = os.path.join(path,f)
        bc,s = _read_bartag(f_name)
        name,ext = os.path.splitext(f)
        first_code = bc[0]
        if mode == 'rename':
            print(first_code)
            shutil.copyfile(f_name,os.path.join(result_path,first_code+ext))
        elif mode == 'print':
            print(first_code)

@register_feature(description="""根据条码信息分割pdf
""")
def split_pdf_on_barcode():
    # 加载全部为图片
    fpath = input('请输入pdf文件路径：').strip('\"').strip('\\')
    with open(fpath, 'rb') as fp:
        pdfFileObject = fp
        pdfReader = PyPDF2.PdfFileReader(pdfFileObject, strict=True)
        pagecount = pdfReader.numPages
        print("总页数：", pagecount)

        root,file = os.path.split(fpath)
        #加载全部为
        print('加载图片中...')
        temp_dir = os.path.join(root,'temp-asegvarefd')
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
        os.mkdir(temp_dir)
        images = convert_from_path(fpath, dpi=300,output_folder=temp_dir,thread_count=8)
        pages_data = []
        for i,im in enumerate(images):
            print(f'识别第{i+1}页图片条码...')
            barcodes, texts = _decode_img(im,barcode_mode=True,text_mode=False)
            pages_data.append((barcodes, texts))
        print('识别完成。')

        splitter = []
        max_ = len(pages_data[0][0])
        for i in range(pagecount):
            if i == 0:
                left = right = 0
                continue
            i1 = len(pages_data[i-1][0])
            i2 = len(pages_data[i][0])
            max_ = max(max_,i2)
            if (i1 == i2) and (pages_data[i-1][0][i1-1] == pages_data[i][0][i2-1]):
                right = i
                if right == pagecount - 1:
                    splitter.append((left,right))
            else:
                if (i2 < max_) and (i1 == i2):

                    right = i
                else:
                    splitter.append((left,right))
                    left = i
                if right == pagecount - 1:
                    splitter.append((left,right))
        _splitter_(splitter,pdfReader,fpath,pages_data)
    shutil.rmtree(temp_dir)

@register_feature(description="""通过表格分割pdf
""")
def split_pdf_on_excel():
    pdfpath = input('输入pdf路径：').strip('\"')
    fpath = input('输入excel表格路径（A、B列必须为左右页码，标题为f,t）：').strip('\"')
    df = pd.read_excel(fpath)
    df.columns = [c.lower() for c in df.columns]
    splitter = []
    for index,row in df.iterrows():
        splitter.append((int(row['f'])-1,int(row['t'])-1))
    with open(pdfpath, 'rb') as fp:
        pdfFileObject = fp
        pdfReader = PyPDF2.PdfFileReader(pdfFileObject, strict=True)
        _splitter_(splitter,pdfReader,pdfpath)

def run():
    try:
        show_features()
        opt = input("输入要运行的条码功能编号（输入help查看完整功能描述）：")
        if opt == "help":
            show_description()
        elif opt.isalnum() and int(opt) < len(features):
            run_feature(features[features_id[int(opt)]])
        else:
            print("无法识别此功能。")
    except Exception as err:
        print(err)
        traceback.print_exc()

if __name__ == '__main__':
    run()