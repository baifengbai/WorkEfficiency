# -*- coding: utf-8 -*-
# @Time    : 9/2/2021 8:36 PM
# @Author  : Chris.Wang
# @Site    : 
# @File    : qiniu_pic_host.py
# @Software: PyCharm
# @Description:将本地图片转为url，使用临时对象存储七牛云（一个月）
import copy
import json
import os
import shutil

from typing import Any, List, AnyStr

from datetime import datetime

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

import PIL.Image

import pandas as pd
import requests
from qiniu import Auth, put_file, etag, build_batch_delete
from qiniu import BucketManager

import filetype

current_path = os.path.dirname(__file__)
with open(os.path.join(current_path, 'qiniu_auth.json'), 'r', encoding='UTF-8') as f:
    auth_info = json.load(f)
    if not auth_info:
        exit('load auth_info failed.')


class QnObS():

    def __init__(self, auth_info):
        self.access_key = auth_info['access_key']
        self.secret_key = auth_info['secret_key']

        self.bucket_name = auth_info['bucket_name']
        self.bucket_domain = None

        self.time_out = 3600
        self.q = None

    def connect(self):
        self.q = Auth(self.access_key, self.secret_key)
        self.bucket_domain = self.get_bucket_domain(self.bucket_name)

    # 上传图片
    def update(self, key, localfile, bucket_name='default_bucket') -> (bool, str):
        if bucket_name == 'default_bucket':
            bucket_name = self.bucket_name
        root, filename = os.path.split(localfile)
        print(f"正在上传：{filename} -> {key}", end='')
        token = self.q.upload_token(bucket_name, key, self.time_out)
        ret, info = put_file(token, key, localfile, version='v1')
        if (ret['key'] == key) and (ret['hash'] == etag(localfile)):
            url = '/'.join([self.bucket_domain, key])
            print('\t上传完成。')
            return True, url
        else:
            url = ''
            return False, url

    def get_file_info(self, key, bucket_name='default_bucket') -> dict:
        if bucket_name == 'default_bucket':
            bucket_name = self.bucket_name
        bucket = BucketManager(self.q)
        ret, info = bucket.stat(bucket_name, key)
        return ret

    # 请求获取空间信息
    def get_bucket_info(self, bucket_name):
        access_token = self.q.token_of_request(f"https://uc.qbox.me/v2/domains?tbl={bucket_name}")
        headers = {
            'Content-Type': "application/x-www-form-urlencoded",
            'Authorization': f'QBox {access_token}'
        }
        response = requests.get(f"https://uc.qbox.me/v2/domains?tbl={bucket_name}", headers=headers)
        return response

    def get_bucket_domain(self, bucket_name):
        ret = self.get_bucket_info(bucket_name)
        return json.loads(ret.text)[0]

    def get_key_url(self, key, bucket_name='default_bucket'):
        if bucket_name == "default_bucket":
            bucket_domain = self.bucket_domain
        else:
            bucket_domain = self.get_bucket_domain(bucket_name)

        return f"http://{bucket_domain}/{key}"

    # localfile to url
    def localfile_to_url(self, root, fpath) -> (bool, str):
        key = fpath.replace(root, '').lstrip('\\').lstrip('/').replace('\\', '/')
        timestamp = datetime.strftime(datetime.now(), '%Y%m%d')
        key = '/'.join([timestamp, key])
        return self.update(key, localfile=fpath)

    def list_keys(self, prefix='', limit=None):
        bucket = BucketManager(self.q)
        delimiter = None
        marker = None
        ret, eof, info = bucket.list(self.bucket_name, prefix, marker, limit, delimiter)
        filelist = [item['key'] for item in ret['items']]
        return filelist

    @staticmethod
    def _del(bucket_inst, bucket_name, keys):
        ops = build_batch_delete(bucket_name, keys)
        print('Will delete: ')
        for key in keys:
            print(key)
        desicion = input('Are you sure? Y or N.')
        if desicion.lower() not in ['y', 'yes']:
            exit('取消了操作')
        ret, info = bucket_inst.batch(ops)
        okcount = [r['code'] for r in json.loads(info.text_body)].count(200)
        print(f"{okcount} files were deleted.")

    def delete_by_keys(self, keys):
        bucket = BucketManager(self.q)
        self._del(bucket, self.bucket_name, keys)

    def delete_by_prefix(self, prefix):
        bucket = BucketManager(self.q)
        limit = None
        delimiter = None
        marker = None
        ret, eof, info = bucket.list(self.bucket_name, prefix, marker, limit, delimiter)
        delete_keys = [item['key'] for item in ret['items']]
        self._del(bucket, self.bucket_name, delete_keys)


# 根据文件目录生成文件路径列表，并且在文件夹内生成excel。
def localfile_to_excel(dirpath) -> str:
    result = os.walk(dirpath)
    items = []
    for root, dirs, files in result:
        for file in files:
            if filetype.guess_extension(os.path.join(root, file)) in ['png', 'jpg', 'gif']:
                items.append((root, file))
    df = pd.DataFrame([(os.path.dirname(dirpath), os.path.join(root, file)) for root, file in items], columns=['root', 'filepath'])
    df['url'] = ''
    excel_path = os.path.join(dirpath, 'images.xlsx')
    df.to_excel(excel_path, engine='openpyxl')

    wb = load_workbook(excel_path)
    ws = wb.active
    # 设定字体
    for col in ws.columns:
        for cell in col:
            cell.font = Font(name="Calibri")
            cell.alignment = Alignment(horizontal='left', wrapText=True)

    for cell in ws[1]:
        font1 = copy.copy(cell.font)
        font1.bold = True
        cell.font = font1

    wb.save(excel_path)
    return excel_path


# 直接对表格进行操作，生成url
def generate_url(excel_path,update=True):
    qos = QnObS(auth_info)
    qos.connect()
    root_path,_ = os.path.split(excel_path)

    wb = load_workbook(excel_path)
    ws = wb.active
    ws['A1'].value = 'img'
    # 生成表头名称字典
    col_names = {}
    current = 1
    for col in ws.iter_cols(1, ws.max_column):
        col_names[col[0].value] = get_column_letter(current)
        current += 1
    temp_dir = os.path.join(root_path,'tempimg-'+datetime.strftime(datetime.now(),'%Y%m%d'))
    if not os.path.exists(temp_dir):
        os.mkdir(temp_dir)
    ws.column_dimensions[col_names['img']].width = 13  # img
    filelist = ws[col_names['filepath']]
    filelist_forup = []
    filelist_forup.append(filelist[0])
    lenfiles = len(filelist)
    blank_img = PIL.Image.new('RGB', (100, 100), color=(255, 255, 255))
    for x in range(1, lenfiles):
        fpath = filelist[x].value
        if fpath.startswith('http'):
            _,temp_name = os.path.split(fpath)
            tempfpath = os.path.join(temp_dir,temp_name)
            retry_count= 0
            while True:
                try:
                    if os.path.exists(tempfpath):
                        break
                    response = requests.get(fpath)
                    if response.status_code == 200:
                        with open(tempfpath, 'wb') as tempf:
                            tempf.write(response.content)
                        print(f'Downloaded {temp_name} successfully.')
                        break
                except:
                    retry_count += 1
                    if retry_count >= 4:
                        print(f'Download ERROR of {temp_name}.')
                        break
            fpath = tempfpath
        filelist_forup.append(fpath)
        ws.row_dimensions[x + 1].height = 70
        if os.path.exists(fpath):
            img = Image(fpath)
        else:
            img = Image(blank_img)
        width = img.width
        height = img.height
        if width >= height:
            img.width = 90
            img.height = img.width / width * height
        else:
            img.height = 90
            img.width = img.height / height * width

        img.anchor = f'A{x + 1}'
        ws.add_image(img)

    if update:
        filepath_list = filelist_forup
        root_list = ws[col_names['root']]
        lenfp = len(filepath_list)
        url_list = ws[col_names['url']]
        for x in range(1, lenfp):
            fpath = filepath_list[x]
            root = root_list[x].value
            result, url = qos.localfile_to_url(root=root, fpath=fpath)
            if result:
                url_list[x].value = 'http://' + url
    # shutil.rmtree(temp_dir)
    wb.save(excel_path)


def del_source(feed, mode='key'):
    opts = ['key', 'keys', 'prefix']
    if mode not in opts:
        exit('del无法识别的模式')
    qos = QnObS(auth_info)
    qos.connect()
    if (mode == 'key') and (isinstance(feed, str)):
        qos.delete_by_keys([feed])
    if (mode == 'keys') and (isinstance(feed, List)):
        qos.delete_by_keys(feed)
    elif (mode == 'prefix') and (isinstance(feed, str)):
        qos.delete_by_prefix(feed)


def list_files(prefix='', mode='dir'):
    if mode not in ['dir', 'all', 'group']:
        exit('list无法识别的模式')
    qos = QnObS(auth_info)
    qos.connect()
    filelist = qos.list_keys(prefix=prefix)
    sliced_filelist = [os.path.split(file) for file in filelist]
    tube = {}
    for root, file in sliced_filelist:
        if root not in tube.keys():
            tube[root] = []
        tube[root].append(file)

    if mode == 'dir':
        temp_root = [key.split('/') for key in tube.keys()]
        set_root = set([t[0] for t in temp_root])
        for key in set_root:
            print(key)
    elif mode == 'all':
        for file in filelist:
            print(file)
    elif mode == 'group':
        for key, value in tube.items():
            print(f"{key}: {', '.join(value)}")


def run():
    def path_mode():
        dirpath = input("输入文件夹路径：").rstrip('\\').strip('\"')
        excel_path = localfile_to_excel(dirpath)
        generate_url(excel_path)

    def excel_mode():
        excelpath = input("输入表格路径：").rstrip('\\').strip('\"')
        generate_url(excelpath)
    def excel_mode2():
        excelpath = input("输入表格路径：").rstrip('\\').strip('\"')
        generate_url(excelpath,update=False)

    def del_mode():
        mode = input("选择文件删除模式（key, keys, prefix）：")
        if mode not in ['key', 'keys', 'prefix']:
            exit('无法识别的模式。')
        if mode == 'prefix':
            feed = input('输入文件前缀：')
            del_source(feed, mode)
        elif mode == 'key':
            feed = input('输入文件名：')
            del_source(feed, mode)
        elif mode == 'keys':
            while True:
                filepath = input("输入文件名表格路径：").strip('\"').strip('\\')
                try:
                    df = pd.read_excel(filepath)
                    break
                except:
                    exit('无法识别的表格文件。')
            feed = df['keys'].to_list()
            del_source(feed, mode)

    def list_mode():
        prefix = input('请输入前缀（默认为空）：')
        show_mode = input('选择显示模式，默认为dir（dir, all, group）：')
        list_files(prefix, mode=show_mode)

    mode_opts = {
        'bydir': path_mode, # add pic to sheet and update by dir
        'byexcel': excel_mode, # add pic to sheet and update
        'byexcel2':excel_mode2, # only add to sheet pic not update
        'del': del_mode,
        'list': list_mode
    }
    mode = input(f'请选择模式 ({", ".join(mode_opts.keys())})：')
    if mode in mode_opts.keys():
        mode_opts[mode]()
    else:
        exit('无法识别的模式。')


if __name__ == '__main__':
    run()
